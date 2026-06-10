import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io
from models import Venta, Compra, Producto

class ReportGenerator:
    @staticmethod
    def _apply_premium_styling(ws, header_fill_color="0F766E"):
        """
        Aplica un diseño ejecutivo de primer nivel a la hoja de cálculo.
        """
        # Definición de estilos
        font_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        fill_header = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type='solid')
        font_data = Font(name='Segoe UI', size=10)
        
        border_thin = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )
        
        border_header = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='medium', color=header_fill_color),
            bottom=Side(style='medium', color=header_fill_color)
        )

        # Forzar visualización de líneas de cuadrícula en Excel
        if ws.views.sheetView:
            ws.views.sheetView[0].showGridLines = True
        else:
            ws.sheet_view.showGridLines = True

        # 1. Altura y estilos para la fila de cabecera
        ws.row_dimensions[1].height = 28
        for cell in ws[1]:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            cell.border = border_header

        # 2. Iterar sobre los datos para dar formato dinámico
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 20
            for cell in ws[row_idx]:
                cell.font = font_data
                cell.border = border_thin
                cell.alignment = Alignment(vertical='center')

                val = cell.value
                if val is None:
                    continue

                # Determinar el nombre de la columna para aplicar formato financiero/numérico
                col_name = ws.cell(row=1, column=cell.column).value
                
                # Formatear números
                if isinstance(val, (int, float)):
                    if col_name in ['Total', 'Precio Venta', 'Costo Unitario', 'Subtotal', 'Monto Total']:
                        # Moneda de Nicaragua (Córdobas)
                        cell.number_format = '"C$"#,##0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    elif col_name in ['Cantidad', 'Stock Actual', 'Mínimo', 'Variación']:
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                
                # Alinear strings técnicos
                else:
                    val_str = str(val)
                    if val_str.startswith('FAC-') or val_str.startswith('COM-') or val_str.startswith('REC-') or val_str.startswith('REF-'):
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif '/' in val_str and len(val_str) <= 19:  # Fechas
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif val_str in ['ENTRADA', 'SALIDA', 'CRÍTICO', 'ÓPTIMO', 'Completada', 'Anulada', 'ENT - COMPRA', 'ENT - PRODUCCIÓN', 'SAL - VENTA', 'SAL - MERMA', 'ENT - ANULACIÓN', 'SAL - ANULACIÓN']:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Resaltado visual para estados
                        if val_str in ['CRÍTICO', 'SAL - MERMA']:
                            cell.font = Font(name='Segoe UI', size=10, bold=True, color='991B1B')
                            cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                        elif val_str in ['ÓPTIMO', 'Completada', 'ENT - COMPRA', 'ENT - PRODUCCIÓN']:
                            cell.font = Font(name='Segoe UI', size=10, bold=True, color='065F46')
                            cell.fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
                        elif val_str in ['Anulada', 'ENT - ANULACIÓN', 'SAL - ANULACIÓN']:
                            cell.font = Font(name='Segoe UI', size=10, italic=True, color='6B7280')
                            cell.fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')

        # 3. Auto-ajuste inteligente de columnas con padding seguro
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = cell.value
                if val is not None:
                    # Si tiene formato de moneda C$ o decimales, agregar un colchón de caracteres adicionales
                    col_name = ws.cell(row=1, column=cell.column).value
                    val_str = str(val)
                    if col_name in ['Total', 'Precio Venta', 'Costo Unitario', 'Subtotal', 'Monto Total'] and isinstance(val, (int, float)):
                        val_str = f"C$ {val_str}.00"
                    
                    if len(val_str) > max_len:
                        max_len = len(val_str)
            # Aplicar ancho dinámico con padding de seguridad
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    @staticmethod
    def generate_sales_excel(ventas):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ventas IVVI"
        
        headers = ['ID Factura', 'Fecha', 'Cliente', 'RUC Cliente', 'Vendedor', 'Monto Total']
        ws.append(headers)
        
        for v in ventas:
            # Si la venta está anulada, agregamos el texto del estado al total
            ws.append([
                f"FAC-{v.id:06d}",
                v.fecha.strftime('%d/%m/%Y %H:%M'),
                v.cliente.nombre if v.cliente else 'N/A',
                v.cliente.ruc if v.cliente else 'N/A',
                v.usuario.nombre if v.usuario else 'Sistema',
                v.total if v.estado != 'Anulada' else 0.0 # O reflejar el total
            ])
            
            # Si está anulada, podemos opcionalmente pintar la celda de estado
            if v.estado == 'Anulada':
                row_idx = ws.max_row
                ws.cell(row=row_idx, column=1).value = f"FAC-{v.id:06d} (ANULADA)"
        
        ReportGenerator._apply_premium_styling(ws, header_fill_color="0F766E")
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_purchases_excel(compras):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Compras IVVI"
        
        headers = ['ID Compra', 'Fecha', 'Factura Proveedor', 'Proveedor', 'Monto Total']
        ws.append(headers)
        
        for c in compras:
            ws.append([
                f"COM-{c.id:06d}",
                c.fecha.strftime('%d/%m/%Y %H:%M'),
                c.numero_factura or 'Sin Ref',
                c.proveedor.razon_social if c.proveedor else 'N/A',
                c.total if c.estado != 'Anulada' else 0.0
            ])
            if c.estado == 'Anulada':
                row_idx = ws.max_row
                ws.cell(row=row_idx, column=1).value = f"COM-{c.id:06d} (ANULADA)"
            
        ReportGenerator._apply_premium_styling(ws, header_fill_color="0F766E")
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_inventory_excel(productos):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario IVVI"
        
        headers = ['SKU', 'Producto', 'Categoría', 'Unidad', 'Precio Venta', 'Stock Actual', 'Mínimo', 'Estado Stock']
        ws.append(headers)
        
        for p in productos:
            estado = "CRÍTICO" if p.stock_actual <= p.stock_minimo else "ÓPTIMO"
            ws.append([
                p.sku,
                p.nombre,
                p.categoria.nombre if p.categoria else 'N/A',
                p.unidad.nombre if p.unidad else 'N/A',
                p.precio_venta,
                p.stock_actual,
                p.stock_minimo,
                estado
            ])
            
        ReportGenerator._apply_premium_styling(ws, header_fill_color="0F766E")
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_kardex_excel(kardex_items):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kardex IVVI"
        
        headers = ['Fecha', 'Hora', 'Tipo Movimiento', 'Documento/REF', 'SKU', 'Producto/Material', 'Cantidad', 'Responsable', 'Observaciones']
        ws.append(headers)
        
        for m in kardex_items:
            sign = '+' if m.tipo_movimiento == 'ENTRADA' else '-'
            
            # Sub-clasificación visual en excel similar al HTML
            doc_str = str(m.documento_id)
            tipo_mov = m.tipo_movimiento
            if m.tipo_movimiento == 'ENTRADA':
                if doc_str.startswith('COMPRA'):
                    tipo_mov = 'ENT - COMPRA'
                elif doc_str.startswith('PLANTA'):
                    tipo_mov = 'ENT - PRODUCCIÓN'
                elif 'AJUSTE-VENTA' in doc_str:
                    tipo_mov = 'ENT - ANULACIÓN'
            else:
                if doc_str.startswith('VENTA'):
                    tipo_mov = 'SAL - VENTA'
                elif 'AJUSTE-COMPRA' in doc_str:
                    tipo_mov = 'SAL - ANULACIÓN'
                elif doc_str.startswith('AJUSTE'):
                    tipo_mov = 'SAL - MERMA'

            ws.append([
                m.fecha.strftime('%d/%m/%Y'),
                m.fecha.strftime('%H:%M:%S'),
                tipo_mov,
                f"REF-{m.documento_id}",
                m.producto.sku if m.producto else 'N/A',
                m.producto.nombre if m.producto else 'N/A',
                m.cantidad if m.tipo_movimiento == 'ENTRADA' else -m.cantidad,
                m.usuario.nombre if m.usuario else 'Sistema',
                m.observacion or 'Procesado por sistema'
            ])
            
        ReportGenerator._apply_premium_styling(ws, header_fill_color="0F766E")
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
